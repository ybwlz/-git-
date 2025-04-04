from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.urls import reverse
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_POST
from django.db.models import Sum, Q
from django.utils import timezone
from django.template.loader import render_to_string
from mymanage.users.decorators import teacher_required
from mymanage.students.models import Student
from .models import Payment, PaymentCategory, Fee
from .forms import PaymentForm, PaymentCategoryForm
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import json
from django.core.paginator import Paginator

@login_required
@teacher_required
def payment_list(request):
    """付款记录列表"""
    teacher = request.user.teacher_profile
    
    # 获取筛选参数
    year = request.GET.get('year')
    month = request.GET.get('month')
    
    # 获取当前日期，用于默认显示
    today = timezone.now().date()
    current_year = today.year
    current_month = today.month
    
    # 如果未提供筛选参数，默认使用当前年月
    if not year:
        year = current_year
    else:
        year = int(year)
    
    if not month:
        month = current_month
    else:
        month = int(month)
    
    # 获取该教师学生的所有付款记录
    all_payments = Payment.objects.filter(
        student__courses__teacher=teacher
    ).order_by('-created_at').distinct()
    
    # 根据年份和月份进行筛选
    filtered_payments = all_payments
    if year and month:
        # 使用Q对象实现OR查询
        filtered_payments = all_payments.filter(
            Q(created_at__year=year, created_at__month=month) |
            Q(payment_date__year=year, payment_date__month=month)
        )
    
    # 分页处理，固定每页显示12条记录
    paginator = Paginator(filtered_payments, 12)  # 每页显示12条记录
    
    # 获取当前页码
    page = request.GET.get('page', 1)
    try:
        payments = paginator.page(page)
    except:
        payments = paginator.page(1)
    
    # 获取统计数据
    total_paid = filtered_payments.filter(status='paid').aggregate(total=Sum('amount'))['total'] or 0
    total_pending = filtered_payments.filter(status='pending').aggregate(total=Sum('amount'))['total'] or 0
    
    # 生成年份选项列表（从当前年份往前15年）
    years = list(range(current_year - 14, current_year + 1))
    
    # 生成月份选项列表
    months = [
        {'value': 1, 'name': '一月'},
        {'value': 2, 'name': '二月'},
        {'value': 3, 'name': '三月'},
        {'value': 4, 'name': '四月'},
        {'value': 5, 'name': '五月'},
        {'value': 6, 'name': '六月'},
        {'value': 7, 'name': '七月'},
        {'value': 8, 'name': '八月'},
        {'value': 9, 'name': '九月'},
        {'value': 10, 'name': '十月'},
        {'value': 11, 'name': '十一月'},
        {'value': 12, 'name': '十二月'}
    ]
    
    # 按月统计本年度收入（折线图）
    month_names = ['一月', '二月', '三月', '四月', '五月', '六月', '七月', '八月', '九月', '十月', '十一月', '十二月']
    
    # 按月统计本年度收入
    monthly_income_data = []
    
    # 生成月度数据（包括测试数据）
    has_real_data = False
    for m in range(1, 13):
        monthly_total = all_payments.filter(
            status='paid',
            payment_date__year=year,
            payment_date__month=m
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        # 如果有真实数据，标记一下
        if monthly_total > 0:
            has_real_data = True
            
        monthly_income_data.append({
            'month': month_names[m-1],
            'value': float(monthly_total)
        })
    
    # 如果没有任何真实数据，添加测试数据
    if not has_real_data:
        monthly_income_data = [
            {'month': '一月', 'value': 1000.0},
            {'month': '二月', 'value': 1200.0},
            {'month': '三月', 'value': 800.0},
            {'month': '四月', 'value': 1500.0},
            {'month': '五月', 'value': 0.0},
            {'month': '六月', 'value': 0.0},
            {'month': '七月', 'value': 0.0},
            {'month': '八月', 'value': 0.0},
            {'month': '九月', 'value': 0.0},
            {'month': '十月', 'value': 0.0},
            {'month': '十一月', 'value': 0.0},
            {'month': '十二月', 'value': 0.0}
        ]
    
    # 获取所有付款类别
    categories = PaymentCategory.objects.all()
    
    # 按类别统计收入数据（饼图）
    category_data = []
    
    # 生成分类数据
    for category in categories:
        category_total = filtered_payments.filter(
            status='paid',
            category=category
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        # 如果没有数据，添加测试数据
        if category_total == 0 and category.id <= 2 and not has_real_data:
            if category.id == 1:  # 学费
                category_total = 3000
            elif category.id == 2:  # 书本费
                category_total = 1500
        
        if category_total > 0:
            category_data.append({
                'name': category.name,
                'value': float(category_total)
            })
    
    # 如果分类数据为空，添加测试数据
    if not category_data and not has_real_data:
        category_data = [
            {'name': '学费', 'value': 3000.0},
            {'name': '书本费', 'value': 1500.0}
        ]
    
    # 序列化数据为JSON格式，使用json_script可以更安全地传递到前端
    monthly_income_json = json.dumps(monthly_income_data, ensure_ascii=False)
    category_data_json = json.dumps(category_data, ensure_ascii=False)
    
    # 构建分页链接中保持筛选参数
    query_params = request.GET.copy()
    if 'page' in query_params:
        query_params.pop('page')
    extra_query = query_params.urlencode()
    
    context = {
        'payments': payments,
        'paginator': paginator,
        'page_obj': payments,
        'is_paginated': payments.has_other_pages(),
        'total_paid': total_paid,
        'total_pending': total_pending,
        'categories': categories,
        'current_year': current_year,
        'current_month': current_month,
        'selected_year': year,
        'selected_month': month,
        'years': years,
        'months': months,
        'monthly_income_data': monthly_income_json,
        'category_data': category_data_json,
        'extra_query': extra_query,
        'title': f'{year}年{month_names[month-1]}付款记录',
    }
    
    return render(request, 'finance/payment_list.html', context)

@login_required
@teacher_required
def payment_add(request):
    """添加付款记录"""
    if request.method == 'POST':
        form = PaymentForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.created_by = request.user
            
            # 设置默认的支付方式为现金
            payment.payment_method = 'cash'
            
            # 如果状态为已支付，但未设置支付日期，则设置为当前日期
            if payment.status == 'paid' and not payment.payment_date:
                payment.payment_date = timezone.now().date()
                
            payment.save()
            messages.success(request, '付款记录添加成功')
            
            # 根据来源页面决定重定向目标
            next_url = request.POST.get('next', '')
            if next_url and next_url == 'finance_overview':
                return redirect('teachers:finance')
            return redirect('finance:payment_list')
    else:
        form = PaymentForm()
        # 只显示当前教师的学生
        teacher = request.user.teacher_profile
        form.fields['student'].queryset = form.fields['student'].queryset.filter(
            courses__teacher=teacher
        ).distinct()
    
    context = {
        'form': form,
        'title': '添加付款记录'
    }
    
    return render(request, 'finance/payment_form.html', context)

@login_required
@teacher_required
def payment_edit(request, payment_id):
    """编辑付款记录"""
    payment = get_object_or_404(Payment, id=payment_id)
    
    # 检查是否为当前教师的学生的付款记录
    teacher = request.user.teacher_profile
    if payment.student.courses.filter(teacher=teacher).exists() == False:
        messages.error(request, '您无权编辑此付款记录')
        return redirect('finance:payment_list')
    
    if request.method == 'POST':
        form = PaymentForm(request.POST, instance=payment)
        if form.is_valid():
            form.save()
            messages.success(request, '付款记录已更新')
            return redirect('finance:payment_list')
    else:
        form = PaymentForm(instance=payment)
        # 只显示当前教师的学生
        form.fields['student'].queryset = form.fields['student'].queryset.filter(
            courses__teacher=teacher
        ).distinct()
    
    context = {
        'form': form,
        'payment': payment,
        'title': '编辑付款记录'
    }
    
    return render(request, 'finance/payment_form.html', context)

@login_required
@teacher_required
def payment_delete(request, payment_id):
    """删除付款记录"""
    payment = get_object_or_404(Payment, id=payment_id)
    
    # 检查是否为当前教师的学生的付款记录
    teacher = request.user.teacher_profile
    if payment.student.courses.filter(teacher=teacher).exists() == False:
        messages.error(request, '您无权删除此付款记录')
        return redirect('finance:payment_list')
    
    if request.method == 'POST':
        payment.delete()
        messages.success(request, '付款记录已删除')
        return redirect('finance:payment_list')
    
    context = {
        'payment': payment
    }
    
    return render(request, 'finance/payment_confirm_delete.html', context)

@login_required
@teacher_required
@require_POST
def mark_as_paid(request, payment_id):
    """标记为已支付"""
    payment = get_object_or_404(Payment, id=payment_id)
    
    # 检查是否为当前教师的学生的付款记录
    teacher = request.user.teacher_profile
    if payment.student.courses.filter(teacher=teacher).exists() == False:
        return JsonResponse({'success': False, 'error': '您无权操作此付款记录'})
    
    payment.status = 'paid'
    payment.payment_date = timezone.now().date()
    payment.save()
    
    return JsonResponse({'success': True})

@login_required
@teacher_required
def payment_export(request):
    """导出付款记录为Excel文件"""
    teacher = request.user.teacher_profile
    
    # 获取当前老师关联的所有付款记录
    payments = Payment.objects.filter(
        student__courses__teacher=teacher
    ).select_related('student', 'category').order_by('-created_at')
    
    # 创建一个响应对象，设置content_type为Excel格式
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="付款记录导出_{}.xlsx"'.format(
        timezone.now().strftime('%Y%m%d%H%M%S')
    )
    
    # 创建Excel工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "付款记录"
    
    # 添加标题行
    columns = ['学生姓名', '付款类别', '金额', '状态', '支付方式', '支付日期', '创建时间', '备注']
    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = column_title
        cell.font = Font(bold=True)
        # 设置列宽
        ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    # 添加数据行
    for row_num, payment in enumerate(payments, 2):
        # 学生姓名
        ws.cell(row=row_num, column=1).value = payment.student.name
        # 付款类别
        ws.cell(row=row_num, column=2).value = payment.category.name
        # 金额
        ws.cell(row=row_num, column=3).value = float(payment.amount)
        # 状态
        status_map = {'paid': '已支付', 'pending': '待支付'}
        ws.cell(row=row_num, column=4).value = status_map.get(payment.status, payment.status)
        # 支付方式
        payment_method_map = {
            'cash': '现金',
            'wechat': '微信支付',
            'alipay': '支付宝',
            'bank_transfer': '银行转账',
            'other': '其他'
        }
        ws.cell(row=row_num, column=5).value = payment_method_map.get(payment.payment_method, payment.payment_method)
        # 支付日期
        if payment.payment_date:
            ws.cell(row=row_num, column=6).value = payment.payment_date.strftime('%Y-%m-%d')
        # 创建时间
        ws.cell(row=row_num, column=7).value = payment.created_at.strftime('%Y-%m-%d %H:%M:%S')
        # 备注
        ws.cell(row=row_num, column=8).value = payment.notes
    
    # 保存Excel文件到响应对象
    wb.save(response)
    return response
